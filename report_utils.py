import datetime
import io

import numpy as np
import openpyxl
from fpdf import FPDF
from openpyxl.drawing.image import Image as XLImage


def build_excel(fig, stats, cap, column, chart_type, subgroup_size, USL, LSL):
    buf_img = io.BytesIO()
    fig.savefig(buf_img, format="png", dpi=150, bbox_inches="tight")
    buf_img.seek(0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SPC Report"

    rows = [
        ["Column", column],
        ["Chart Type", chart_type],
        ["Subgroup Size", subgroup_size],
        ["CL (X-bar)", round(stats["xbar_bar"], 4)],
        ["UCL", round(stats["UCL_xbar"], 4)],
        ["LCL", round(stats["LCL_xbar"], 4)],
        ["USL", USL],
        ["LSL", LSL],
        ["Sigma (ST)", round(cap["sigma_st"], 4) if cap["sigma_st"] is not None else "N/A"],
        ["Sigma (LT)", round(cap["sigma_lt"], 4) if cap["sigma_lt"] is not None else "N/A"],
        ["Cp",  round(cap["Cp"],  4) if cap["Cp"]  is not None else "N/A"],
        ["Cpk", round(cap["Cpk"], 4) if cap["Cpk"] is not None else "N/A"],
        ["Pp",  round(cap["Pp"],  4) if cap["Pp"]  is not None else "N/A"],
        ["Ppk", round(cap["Ppk"], 4) if cap["Ppk"] is not None else "N/A"],
    ]
    for row in rows:
        ws.append(row)

    ws2 = wb.create_sheet("Data")
    ws2.append(["Subgroup", "X-bar", stats["sub_label"]])
    for i, (xb, ss) in enumerate(zip(stats["xbar"], stats["sub_stat"])):
        ws2.append([
            i + 1,
            round(float(xb), 4),
            round(float(ss), 4) if not np.isnan(ss) else "",
        ])

    img = XLImage(buf_img)
    img.anchor = "D1"
    ws.add_image(img)

    buf_out = io.BytesIO()
    wb.save(buf_out)
    buf_out.seek(0)
    return buf_out


def build_pdf(fig, stats, cap, column, chart_type, subgroup_size, USL, LSL, values, inferences):
    buf_img = io.BytesIO()
    fig.savefig(buf_img, format="png", dpi=150, bbox_inches="tight")
    buf_img.seek(0)

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_margins(15, 15, 15)
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()

    page_w = pdf.w - 30  # usable width: A4 210mm minus 15mm margins each side

    # Header
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "SPC Control Chart Report", new_x="LMARGIN", new_y="NEXT", align="C")

    _chart_label = {
        "xbar_r": "X-bar / R Chart",
        "xbar_s": "X-bar / S Chart",
        "im_r":   "Individuals / Moving Range (I-MR) Chart",
    }
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(
        0, 6,
        f"Column: {column}   |   Chart: {_chart_label.get(chart_type, chart_type)}   |   Date: {datetime.date.today()}",
        new_x="LMARGIN", new_y="NEXT", align="C",
    )
    pdf.ln(3)

    # Chart image — full usable width
    pdf.image(buf_img, x=15, w=page_w)
    pdf.ln(4)

    # Stats table — 3 equal columns
    def _fmt(v):
        return "N/A" if v is None else f"{v:.4f}"

    col_w   = page_w / 3
    row_h   = 5
    label_w = col_w * 0.60
    val_w   = col_w * 0.40

    sections = [
        ("Capability", [
            ("Cp",  _fmt(cap["Cp"])),
            ("Cpk", _fmt(cap["Cpk"])),
            ("Pp",  _fmt(cap["Pp"])),
            ("Ppk", _fmt(cap["Ppk"])),
        ]),
        ("Process", [
            ("Mean (X-bar)", f"{stats['xbar_bar']:.4f}"),
            ("Sigma ST",     _fmt(cap["sigma_st"])),
            ("Sigma LT",     _fmt(cap["sigma_lt"])),
            ("Min",          f"{values.min():.4f}"),
            ("Max",          f"{values.max():.4f}"),
            ("Observations", str(len(values))),
            ("Subgroup Size", str(subgroup_size)),
        ]),
        ("Control Limits", [
            ("UCL (X-bar)", f"{stats['UCL_xbar']:.4f}"),
            ("CL (X-bar)",  f"{stats['xbar_bar']:.4f}"),
            ("LCL (X-bar)", f"{stats['LCL_xbar']:.4f}"),
            ("USL",         f"{USL:.4f}"),
            ("LSL",         f"{LSL:.4f}"),
        ]),
    ]

    # Section header row
    y_header = pdf.get_y()
    for i, (title, _) in enumerate(sections):
        pdf.set_xy(15 + i * col_w, y_header)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(col_w, row_h + 1, title, border=1, align="C")
    pdf.set_y(y_header + row_h + 1)

    # Data rows
    max_rows = max(len(rows) for _, rows in sections)
    for row_idx in range(max_rows):
        y = pdf.get_y()
        for i, (_, rows) in enumerate(sections):
            pdf.set_xy(15 + i * col_w, y)
            if row_idx < len(rows):
                label, value = rows[row_idx]
                pdf.set_font("Helvetica", "B", 8)
                pdf.cell(label_w, row_h, label, border=1)
                pdf.set_font("Helvetica", "", 8)
                pdf.cell(val_w, row_h, value, border=1)
            else:
                pdf.cell(col_w, row_h, "", border=1)
        pdf.set_y(y + row_h)

    # Inferences section
    _char_map = {
        "✅": "[OK]",
        "⚠️": "[!]",
        "❌": "[X]",
        "ℹ️": "[i]",
        "—": "-",   # em dash —
        "–": "-",   # en dash –
        "≥": ">=",  # ≥
        "≤": "<=",  # ≤
        "’": "'",   # right single quote '
        "‘": "'",   # left single quote '
        "“": '"',   # left double quote "
        "”": '"',   # right double quote "
    }

    def _pdf_safe(text):
        for char, replacement in _char_map.items():
            text = text.replace(char, replacement)
        # Final fallback: replace any remaining non-latin-1 character
        return text.encode("latin-1", errors="replace").decode("latin-1")


    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "Process Inferences", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(1)

    for b in inferences["bullets"]:
        pdf.set_font("Helvetica", "", 8)
        line = _pdf_safe(f"{b['icon']}  {b['text']}")
        pdf.cell(0, 5, line, new_x="LMARGIN", new_y="NEXT")

    pdf.ln(2)
    pdf.set_fill_color(240, 240, 240)
    pdf.set_font("Helvetica", "I", 8)
    pdf.multi_cell(0, 5, _pdf_safe(inferences["narrative"]), fill=True)
    pdf.ln(2)

    # Footer
    pdf.set_y(pdf.h - 15)
    pdf.set_font("Helvetica", "I", 7)
    pdf.cell(0, 5, "Generated by SPC Control Chart Tool - Internal Use Only", align="R")

    buf_out = io.BytesIO()
    buf_out.write(pdf.output())
    buf_out.seek(0)
    return buf_out
